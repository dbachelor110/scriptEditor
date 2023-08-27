import random
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
from time import sleep
# from Parameter_settings import url, username, password, username_locator, password_locator, submit_locator
import datetime                      #使用日期間距
from openpyxl import load_workbook   #讀excel檔
import ddddocr                       #20230407 for登入驗證
from PIL import Image                #20230407 for登入驗證
import urllib.request                #20230407 for登入驗證
import pyautogui                     #20230412 for上傳檔案
import shutil                        #20230504 for移動檔案位置
from selenium.webdriver.common.keys import Keys     #20230524 for連續動作
from selenium.webdriver.common.action_chains import ActionChains     #20230524 for連續動作
from bs4 import BeautifulSoup        #20230816 for報表
import requests                      #20230816 for報表
import core

# 定義打印指定內容
def print_func(browser, params1=None, params2=None, params3=None,params4=None):
    print(params1)

# 定義產生parms1~parms2之間的亂數並打印的Function
def random_func(browser, params1=None, params2=None, params3=None,params4=None):
    parms1 = int(params1)
    parms2 = int(params2)
    rand_int = random.randint(parms1, parms2)
    print(f"產生{params1}~{params2}中的隨機亂數：{rand_int}")

# 定義 load_params 函数
def load_params(browser=None, param_name=None, params2=None, params3=None,params4=None):
    if f"params_{param_name}" in globals():
        text = globals().get(f"params_{param_name}")
        print(f"Get params_{param_name} = {text}")
        return text
    else:
        print(f"No Find params_{param_name}")

# 定義登入for 和安系統
def login2(browser, params1=None, params2=None, params3=None,params4=None):
    parameterData=core.makeBrowser(core.getParameter())
    browser.get(parameterData['url'])

    # 輸入帳號和密碼
    username_element = WebDriverWait(browser, 10).until(EC.visibility_of_element_located(parameterData['username_locator']))
    username_element.clear()
    username_element.send_keys(parameterData['username'])
    password_element = browser.find_element(*parameterData['password_locator'])
    password_element.clear()
    password_element.send_keys(parameterData['password'])

    #20230407 for登入驗證
    sleep(2)
    img_src = browser.find_element(By.XPATH,'//*[@id="app"]/div/div/div[2]/div[3]/div[2]/button/img').get_attribute('src')
    urllib.request.urlretrieve(img_src, 'image.png')
    img = Image.open('image.png')       # 讀取圖片
    img = img.convert('RGB')
    ocr = ddddocr.DdddOcr()             
    verification_code = ocr.classification(img)     # 使用OCR辨識驗證碼
    browser.find_element(By.XPATH,"//*[@id='app']/div/div/div[2]/div[3]/div[2]/input").send_keys(str.upper(verification_code))

    # 點擊登入按鈕
    submit_element = browser.find_element(*parameterData['submit_locator'])
    submit_element.click()

    # 等待跳轉完成
    try:
        WebDriverWait(browser, 10).until(EC.url_changes(parameterData['url']))
        print('登入成功')
        browser.maximize_window()
    except TimeoutException:
        print('登入失敗')

# 定義 key_in2 行為
# 20230410 新增 (只能用XPATH)
def key_in2(browser, params1=None, params2=None, params3=None,params4=None):
    if params2 != '0':
        text = params2
    elif params3:
        text = load_params(param_name = params3)         #為了使用參數方式
    else:
        print('欄位查不到定義')
    browser.find_element(By.XPATH, params1).send_keys(text)

# 定義 save_as 函数
# 202303xx 新增
def save_as(browser, params1=None, params2=None, params3=None,params4=None):
    value_op = None
    if params1 == 'date_now':
        param_value = datetime.datetime.now().strftime("%Y-%m-%d") 
        globals()[f"params_{params1}"] = param_value
    elif params1 == 'serial_key':
        param_value = datetime.datetime.now().strftime("%m%d%H%M")
        print("以當下時間做變數 : " + param_value)
        globals()[f"params_{params1}"] = param_value
    elif params1 == 'n_date' or params1 == 'n_date2':
        number_day = int (params2)
        day_a = datetime.datetime.strptime(datetime.datetime.now().strftime("%Y-%m-%d"), "%Y-%m-%d") + datetime.timedelta(days=number_day)
        day_y = int(datetime.datetime.now().strftime("%Y"))-1911
        n_date = str(day_y) + str(day_a)[5:7] + str(day_a)[8:10]
        n_date2 = str(day_y) +"/"+ str(day_a)[5:7] +"/"+ str(day_a)[8:10]
        if params1 == 'n_date':
            globals()[f"params_{params1}"] = n_date
        else:
            globals()[f"params_{params1}"] = n_date2
    elif params1 == 'operator':
        op_value1 = globals()[f"params_{params2}"]
        op_value2 = globals()[f"params_{params3}"]
        value_op = float(op_value1.replace(",","")) - float(op_value2.replace(",",""))
        globals()[f"params_{params1}"] = f'{int(value_op):,d}'
    else:
        print('查不到定義')

# 定義 excel_input 函数 20230504
def excel_input(browser, params1=None, params2=None, params3=None,params4=None):
    element = None
    element = browser.find_element(By.XPATH, params1)
    if element:
        # 欄位值可能有兩種形式：text or attribute，會兩種都抓然後塞有值的那一種
        text_value = element.text
        attribute_value = element.get_attribute('value')
        actual_value = text_value if text_value else attribute_value

    # 設定路徑
    excel_path = r"C:\Users\nicK_chang\Desktop\test02.xlsx"
    # 創建新的 DataFrame
    #first_df = pd.DataFrame({'Column A': ['保批單號碼'], 'Column B': ['保單號碼'], 'Column C': ['保險公司代號'], 'Column D': ['保險公司名稱'], 'Column E': ['保險公司代號'], 'Column F': ['保險公司險別名稱'], 'Column G': ['保費'], 'Column H': ['佣金']})
    # 創建 ExcelWriter 物件
    #with pd.ExcelWriter(excel_path) as writer:
    #    first_df.to_excel(writer, sheet_name='Sheet1', header=None, index=False, startrow=0, startcol=0)
    
    # 讀檔
    wb = load_workbook(excel_path)
    ws = wb.active
    for params3 in range(params3, params4+1):
        ws[params2 + str(params3)] = actual_value
    # 儲存檔案
    wb.save(excel_path)

# 定義windows上傳視窗檔案位置
def upload_file(browser, params1,v2,v3,v4):
    print(params1)
    sleep(1)
    pyautogui.typewrite(params1)
    pyautogui.press('enter')
    sleep(1)

# 定義複製檔案至指定路徑
def copy_file(browser, params1=None, params2=None, params3=None,params4=None):
    print(params2)
    shutil.copy2(params1, params2)

# 定義清除欄位值
def clear_func(browser, params1=None, params2=None, params3=None,params4=None):
    browser.find_element(By.XPATH, params1).clear()

# 定義連續按鍵輸入 
def multi_key(browser, params1=None, params2=None, params3=None,params4=None):
    browser.find_element(By.XPATH, params1).click()
    actions = ActionChains(browser)
    actions.send_keys(Keys.DOWN)
    actions.send_keys(Keys.DOWN)
    actions.send_keys(Keys.ENTER)
    actions.perform()

# 定義指定欄位值
def find_word(browser, params1=None, params2=None, params3=None,params4=None):
    content_popup = browser.find_element(By.CSS_SELECTOR, params1).text
    print(content_popup)
    get_word = content_popup[params2:params3+1]
    print("存入get_text變數 : " + get_word)
    globals()[f"params_{params4}"] = get_word

# 定義關閉警告彈窗
def click_alert(browser, params1=None, params2=None, params3=None,params4=None):
    browser.switch_to.alert.accept()

# 定義比對選單或勾選框
def is_selected(browser, params1=None, params2=None, params3=None,params4=None):
    if params3 == '1':
        x = browser.find_element(By.XPATH, params1).is_selected() 
        print (x)
        if x :
            print("比對成功!")
        else:
            print("失敗!!!!!!")
    elif params3 == '2':
        y = browser.find_element(By.XPATH, params1).get_attribute('value')
        print (y)
        if params2 == '0':
            if y == '':
                print("比對成功!")
        else:
            if y == params2:
                print("比對成功!")
            else :
                print("失敗!!!!!!")

# 定義報表比對
def report(browser, params1=None, params2=None, params3=None,params4=None):
    report_url = browser.find_element(By.XPATH, params1).get_attribute('src')
    response = requests.get(report_url)
    soup = BeautifulSoup(response.text, "html.parser")
    #s1 = soup.find_all('span')
    s2 = soup.find_all('td')
    title_report1 = s2[int(params3)]
    print("報表內容: " + title_report1.getText())
    if params2 == title_report1.getText() :
        print("比對成功!")
    else:
        print("失敗!!!!!!")

# 定義連續按鍵輸入2
def multi_key2(browser, params1=None, params2=None, params3=None,params4=None):
    browser.find_element(By.XPATH, params1).click()
    actions = ActionChains(browser)
    actions.send_keys(Keys.TAB)
    actions.send_keys(Keys.ENTER)
    actions.perform()

# 特殊操作
def java_click(browser, params1=None, params2=None, params3=None,params4=None):
    element_b = WebDriverWait(browser, 10).until(EC.element_to_be_clickable((By.XPATH, params1)))
    browser.execute_script("arguments[0].click();", element_b)
